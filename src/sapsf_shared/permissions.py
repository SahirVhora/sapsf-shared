"""SAP SuccessFactors RBP Permission Analyzer.

Wraps the OData v2 Role-Based Permission function imports to pull roles,
permissions, user assignments, and the permission catalogue from a tenant.

Key function imports (all require RBP admin access):
    - getRolesPermissions(locale, roleIds)     → XML: permissions per role
    - getUsersPermissions(locale, userIds)     → XML: permissions per user
    - getUserRolesReport(userIds)              → XML: role assignments per user
    - getUserRolesByUserId(userId)             → list: roles for one user
    - getPermissionMetadata(locale)            → XML: full permission catalogue
    - getUsersByDynamicGroup(groupId, activeOnly) → list: users in a group
    - checkUserPermission(...)                 → bool: quick permission check

Usage:
    config = AuthConfig(base_url="...", username="...", password="...")
    client = SFClient(config)
    analyzer = PermissionAnalyzer(client)
    report = analyzer.full_scan()
"""

from __future__ import annotations

import logging
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Any

from sapsf_shared.client import SFClient
from sapsf_shared.exceptions import SFClientError
from sapsf_shared.utils import odata_escape

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────

_BATCH_LIMIT_USERS = 100
_BATCH_LIMIT_ROLES = 100
_DEFAULT_LOCALE = "en_US"
_RBP_PATH = "odata/v2"
_XML_NS = "{http://www.w3.org/2001/XMLSchema-instance}"
_ROLE_NAME_RE = re.compile(r"<roleName[^>]*>(.*?)</roleName>")
_ROLE_ID_RE = re.compile(r"<roleId[^>]*>(.*?)</roleId>")

# Sensitive permission patterns for risk flagging
_SENSITIVE_PERMISSIONS: dict[str, str] = {
    "MANAGE_COMPENSATION": "Compensation Data Access",
    "VIEW_COMPENSATION": "View Compensation Information",
    "IMPORT_EXPORT_DATA": "Import/Export Data",
    "MANAGE_USER_PROXY": "Proxy Access",
    "MANAGE_PERMISSION_ROLES": "Manage Permission Roles",
    "MANAGE_SECURITY": "Security Administration",
    "MANAGE_EMPLOYEE_DATA": "Manage Employee Data",
    "VIEW_PERSONAL_INFO": "View Personal Information",
    "MANAGE_PAYROLL": "Payroll Administration",
    "MANAGE_BENEFITS": "Benefits Administration",
    "MANAGE_RECRUITING": "Recruiting Administration",
    "MANAGE_SUCCESSION": "Succession Management",
    "MANAGE_PERFORMANCE": "Performance Management Administration",
    "MANAGE_TIME_OFF": "Time Off Administration",
    "API_ACCESS": "API/SFAPI Access",
    "ODATA_API_READ": "OData API Read Access",
    "ODATA_API_WRITE": "OData API Write Access",
}

# Combinations of sensitive permissions that are high-risk
_HIGH_RISK_COMBOS: list[tuple[str, ...]] = [
    ("IMPORT_EXPORT_DATA", "MANAGE_COMPENSATION"),
    ("IMPORT_EXPORT_DATA", "VIEW_PERSONAL_INFO"),
    ("MANAGE_PERMISSION_ROLES", "MANAGE_SECURITY"),
    ("PROXY_ACCESS", "MANAGE_EMPLOYEE_DATA"),
    ("ODATA_API_READ", "IMPORT_EXPORT_DATA"),
    ("ODATA_API_WRITE", "MANAGE_EMPLOYEE_DATA"),
]


# ── Data models ────────────────────────────────────────────────────────────


@dataclass
class PermissionRole:
    """A single permission role with its granted permissions."""

    role_id: str
    role_name: str
    permissions: list[str] = field(default_factory=list)
    permission_categories: dict[str, list[str]] = field(default_factory=dict)
    user_count: int = 0
    is_empty: bool = False


@dataclass
class UserRoleAssignment:
    """A user and the roles assigned to them."""

    user_id: str
    username: str
    full_name: str
    status: str
    role_ids: list[str] = field(default_factory=list)
    role_names: list[str] = field(default_factory=list)
    is_inactive: bool = False


@dataclass
class PermissionCatalogue:
    """The full permission catalogue from getPermissionMetadata."""

    categories: dict[str, list[dict[str, str]]] = field(default_factory=dict)
    raw_xml: str = ""


@dataclass
class PermissionScanReport:
    """Complete scan result."""

    roles: list[PermissionRole]
    users: list[UserRoleAssignment]
    catalogue: PermissionCatalogue | None
    tenant_url: str
    total_roles: int = 0
    total_users: int = 0
    empty_roles: list[PermissionRole] = field(default_factory=list)
    high_risk_roles: list[tuple[str, str, list[str]]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ── XML parsing helpers ────────────────────────────────────────────────────


def _strip_ns(tag: str) -> str:
    """Strip XML namespace from an element tag."""
    return tag.split("}")[-1] if "}" in tag else tag


def _parse_role_permissions_xml(xml_text: str) -> dict[str, list[str]]:
    """Parse the XML response from getRolesPermissions.

    Expected structure (simplified):
        <rolePermissions>
            <role>
                <roleName>...</roleName>
                <roleId>...</roleId>
                <permissions>
                    <permission>PERM_CODE</permission>
                    ...
                </permissions>
            </role>
            ...
        </rolePermissions>
    """
    result: dict[str, list[str]] = {}
    try:
        root = ET.fromstring(xml_text)
        # Handle both namespaced and non-namespaced XML
        for role_el in root.iter():
            tag = _strip_ns(role_el.tag)
            if tag != "role":
                continue

            # Find role name/id
            role_name = ""
            role_id = ""
            key = ""
            for child in role_el.iter():
                ctag = _strip_ns(child.tag)
                if ctag == "roleName":
                    role_name = (child.text or "").strip()
                    key = role_name
                elif ctag == "roleId":
                    role_id = (child.text or "").strip()
                    if not key and role_id:
                        key = role_id
                elif ctag == "permission":
                    perm = (child.text or "").strip()
                    if perm:
                        result.setdefault(key, []).append(perm)

            # Ensure we have at least an entry even if no permissions
            if key:
                result.setdefault(key, [])
    except ET.ParseError as exc:
        logger.warning("Failed to parse role permissions XML: %s", exc)
        # Try regex-based fallback
        return _parse_role_permissions_regex(xml_text)

    return result


def _parse_role_permissions_regex(xml_text: str) -> dict[str, list[str]]:
    """Fallback parser using regex for malformed XML."""
    result: dict[str, list[str]] = {}
    current_role: str | None = None

    # Simple state machine over lines
    for line in xml_text.split("\n"):
        role_name_m = re.search(r"<roleName[^>]*>(.*?)</roleName>", line)
        if role_name_m:
            current_role = role_name_m.group(1).strip()
            if current_role:
                result.setdefault(current_role, [])
            continue
        perm_m = re.search(r"<permission[^>]*>(.*?)</permission>", line)
        if perm_m and current_role:
            result[current_role].append(perm_m.group(1).strip())

    return result


def _parse_user_roles_xml(xml_text: str) -> dict[str, list[str]]:
    """Parse the XML response from getUserRolesReport.

    Expected structure:
        <userRolesReport>
            <userRole>
                <userName>...</userName>
                <roles>
                    <role>ROLE_NAME</role>
                    ...
                </roles>
            </userRole>
            ...
        </userRolesReport>
    """
    result: dict[str, list[str]] = {}
    try:
        root = ET.fromstring(xml_text)
        for user_el in root.iter():
            tag = _strip_ns(user_el.tag)
            if tag != "userRole":
                continue

            username = ""
            roles: list[str] = []
            for child in user_el.iter():
                ctag = _strip_ns(child.tag)
                if ctag == "userName":
                    username = (child.text or "").strip()
                elif ctag == "role":
                    role_name = (child.text or "").strip()
                    if role_name:
                        roles.append(role_name)

            if username:
                result[username] = roles
    except ET.ParseError:
        return _parse_user_roles_regex(xml_text)

    return result


def _parse_user_roles_regex(xml_text: str) -> dict[str, list[str]]:
    """Regex fallback for getUserRolesReport XML."""
    result: dict[str, list[str]] = {}
    current_user: str | None = None

    for line in xml_text.split("\n"):
        user_m = re.search(r"<userName[^>]*>(.*?)</userName>", line)
        if user_m:
            current_user = user_m.group(1).strip()
            result.setdefault(current_user, [])
            continue
        role_m = re.search(r"<role[^>]*>(.*?)</role>", line)
        if role_m and current_user:
            result[current_user].append(role_m.group(1).strip())

    return result


def _parse_users_permissions_xml(xml_text: str) -> dict[str, list[str]]:
    """Parse the XML response from getUsersPermissions.

    Expected structure:
        <usersPermissions>
            <userPermission>
                <userName>...</userName>
                <permissions>
                    <permission>PERM_CODE</permission>
                    ...
                </permissions>
            </userPermission>
            ...
        </usersPermissions>
    """
    result: dict[str, list[str]] = {}
    try:
        root = ET.fromstring(xml_text)
        for up_el in root.iter():
            tag = _strip_ns(up_el.tag)
            if tag != "userPermission":
                continue

            username = ""
            perms: list[str] = []
            for child in up_el.iter():
                ctag = _strip_ns(child.tag)
                if ctag == "userName":
                    username = (child.text or "").strip()
                elif ctag == "permission":
                    perm = (child.text or "").strip()
                    if perm:
                        perms.append(perm)

            if username:
                result[username] = perms
    except ET.ParseError:
        pass

    return result


def _parse_permission_metadata_xml(xml_text: str) -> dict[str, list[dict[str, str]]]:
    """Parse the XML from getPermissionMetadata.

    Expected structure:
        <permissionMetadata>
            <category name="CATEGORY_NAME">
                <permission>
                    <code>PERM_CODE</code>
                    <label>Human Readable Name</label>
                </permission>
                ...
            </category>
            ...
        </permissionMetadata>
    """
    result: dict[str, list[dict[str, str]]] = {}
    try:
        root = ET.fromstring(xml_text)
        for cat_el in root.iter():
            tag = _strip_ns(cat_el.tag)
            if tag == "category":
                cat_name = cat_el.attrib.get("name", "").strip()
                perms: list[dict[str, str]] = []
                for perm_el in cat_el.iter():
                    ptag = _strip_ns(perm_el.tag)
                    if ptag == "permission":
                        code = ""
                        label = ""
                        for child in perm_el.iter():
                            ctag = _strip_ns(child.tag)
                            if ctag == "code":
                                code = (child.text or "").strip()
                            elif ctag == "label":
                                label = (child.text or "").strip()
                        if code:
                            perms.append({"code": code, "label": label or code})
                if cat_name:
                    result[cat_name] = perms
    except ET.ParseError:
        pass

    return result


# ── PermissionAnalyzer ─────────────────────────────────────────────────────


class PermissionAnalyzer:
    """Analyze RBP permissions across an SF tenant.

    Requires the API user to have RBP administrator access
    (granted via Admin Center > Manage Role-Based Permission Access).
    """

    def __init__(
        self,
        client: SFClient,
        locale: str = _DEFAULT_LOCALE,
    ) -> None:
        self._client = client
        self._base_url = client.base_url.replace("/odata/v2", "").rstrip("/")
        self._locale = locale

    # ── Internal: build RBP function URLs ───────────────────────────────

    def _rbp_url(self, function_name: str) -> str:
        """Build URL for an RBP function import."""
        return f"{self._base_url}/odata/v2/{function_name}"

    def _call_rbp_function(
        self,
        function: str,
        params: dict[str, str] | None = None,
        method: str = "GET",
        payload: Any = None,
    ) -> str:
        """Call an RBP function import and return the raw response text.

        These functions return XML, not JSON, so we bypass the normal
        SFClient JSON parsing.
        """
        url = self._rbp_url(function)

        if method == "GET":
            resp = self._client._request_with_retry("GET", url, params=params)
        else:
            resp = self._client._request_with_retry(method, url, json=payload)

        # Handle network-level errors (MagicMock fallback from tests, None responses)
        if resp is None:
            raise SFClientError(
                f"No response from {function} — check network / tenant URL",
                url=url,
            )

        # Handle errors
        if resp.status_code == 401:
            raise SFClientError(
                "Authentication failed — check API user credentials",
                status_code=401,
                url=url,
            )
        if resp.status_code == 403:
            raise SFClientError(
                "Access denied — API user needs RBP administrator access. "
                "Grant via Admin Center > Manage Role-Based Permission Access",
                status_code=403,
                url=url,
            )
        if resp.status_code >= 400:
            raise SFClientError(
                f"HTTP {resp.status_code} from {function}",
                status_code=resp.status_code,
                body=resp.text[:2000],
                url=url,
            )

        return resp.text

    # ── Public API ───────────────────────────────────────────────────────

    def get_permission_metadata(self) -> PermissionCatalogue:
        """Fetch the full permission catalogue.

        This returns every possible permission and its category/label
        for the tenant's locale. Use this to translate permission codes
        to human-readable names.
        """
        xml_text = self._call_rbp_function(
            "getPermissionMetadata",
            params={"locale": self._locale},
        )
        categories = _parse_permission_metadata_xml(xml_text)
        return PermissionCatalogue(categories=categories, raw_xml=xml_text)

    def get_roles_permissions(self, role_ids: list[str]) -> dict[str, list[str]]:
        """Fetch permissions for up to 100 roles.

        Automatically batches if more than 100 roles are provided.
        Returns dict of role_name/role_id → list of permission codes.
        """
        result: dict[str, list[str]] = {}
        for i in range(0, len(role_ids), _BATCH_LIMIT_ROLES):
            batch = role_ids[i : i + _BATCH_LIMIT_ROLES]
            role_ids_param = ",".join(batch)
            try:
                xml_text = self._call_rbp_function(
                    "getRolesPermissions",
                    params={"locale": self._locale, "roleIds": role_ids_param},
                )
                parsed = _parse_role_permissions_xml(xml_text)
                result.update(parsed)
            except SFClientError as exc:
                # Re-raise auth/access errors so callers can detect them
                if exc.status_code == 403:
                    raise
                logger.warning("Failed to get permissions for roles batch: %s", exc)
                # Add empty entries for roles in this batch so we know they exist
                for rid in batch:
                    result.setdefault(rid, [])

        return result

    def get_user_roles_report(
        self, user_ids: list[str]
    ) -> dict[str, list[str]]:
        """Fetch role assignments for up to 100 users.

        Returns dict of username → list of role names.
        Max 100 users per call; automatically batches.
        """
        result: dict[str, list[str]] = {}
        for i in range(0, len(user_ids), _BATCH_LIMIT_USERS):
            batch = user_ids[i : i + _BATCH_LIMIT_USERS]
            user_ids_param = ",".join(batch)
            try:
                xml_text = self._call_rbp_function(
                    "getUserRolesReport",
                    params={"userIds": user_ids_param},
                )
                parsed = _parse_user_roles_xml(xml_text)
                result.update(parsed)
            except SFClientError as exc:
                # Re-raise auth/access errors so callers can detect them
                if exc.status_code == 403:
                    raise
                logger.warning("Failed to get roles for users batch: %s", exc)

        return result

    def get_user_roles_by_user_id(self, user_id: str) -> list[dict[str, Any]]:
        """Fetch role list for a single user.

        Unlike getUserRolesReport, this returns JSON via the standard
        OData client (it returns PermissionRoles entity, not XML).
        """
        try:
            raw = self._call_rbp_function(
                "getUserRolesByUserId",
                params={"userId": user_id},
            )
            # Try parsing as JSON (SF returns JSON for this endpoint)
            try:
                import json as _json
                data = _json.loads(raw)
                results = data.get("d", {}).get("results", [])
                if not results and "d" in data:
                    # Maybe it's a single object, not array
                    single = data.get("d", {})
                    if "roleName" in single:
                        return [single]
                return results
            except (ValueError, TypeError):
                # Fallback: parse XML
                roles = _parse_user_roles_xml(raw)
                return [
                    {"roleName": r, "roleId": r}
                    for r in roles.get(user_id, [])
                ]
        except SFClientError as exc:
            logger.warning("Failed to get roles for user %s: %s", user_id, exc)
            return []

    def get_users_permissions(
        self, user_ids: list[str]
    ) -> dict[str, list[str]]:
        """Fetch effective permissions for up to 100 users.

        Returns dict of username → list of permission codes.
        """
        result: dict[str, list[str]] = {}
        for i in range(0, len(user_ids), _BATCH_LIMIT_USERS):
            batch = user_ids[i : i + _BATCH_LIMIT_USERS]
            user_ids_param = ",".join(batch)
            try:
                xml_text = self._call_rbp_function(
                    "getUsersPermissions",
                    params={"locale": self._locale, "userIds": user_ids_param},
                )
                parsed = _parse_users_permissions_xml(xml_text)
                result.update(parsed)
            except SFClientError as exc:
                logger.warning("Failed to get permissions for users batch: %s", exc)

        return result

    def get_users_by_dynamic_group(
        self, group_id: str, active_only: bool = True
    ) -> list[dict[str, Any]]:
        """Fetch users in a dynamic permission group."""
        try:
            raw = self._call_rbp_function(
                "getUsersByDynamicGroup",
                params={
                    "groupId": str(group_id),
                    "activeOnly": str(active_only).lower(),
                },
            )
            try:
                import json as _json
                data = _json.loads(raw)
                return data.get("d", {}).get("results", [])
            except (ValueError, TypeError):
                return []
        except SFClientError as exc:
            logger.warning(
                "Failed to get users for group %s: %s", group_id, exc
            )
            return []

    def check_user_permission(
        self,
        access_user_id: str,
        perm_type: str,
        perm_string_value: str,
        target_user_id: str | None = None,
    ) -> bool:
        """Check if a user has a specific permission.

        This is useful for targeted checks after the full scan.
        """
        params = {
            "accessUserId": access_user_id,
            "permType": perm_type,
            "permStringValue": perm_string_value,
        }
        if target_user_id:
            params["targetUserId"] = target_user_id

        try:
            raw = self._call_rbp_function("checkUserPermission", params=params)
            return raw.strip().lower() in ("true", "1", "yes")
        except SFClientError:
            return False

    # ── High-level scan ─────────────────────────────────────────────────

    def full_scan(
        self,
        max_users: int = 0,  # 0 = unlimited
        filter_field: str | None = None,
        filter_value: str | None = None,
    ) -> PermissionScanReport:
        """Run a full permission scan against the tenant.

        1. Fetch all active users from the User entity (optionally filtered)
        2. Get role assignments for each user
        3. Discover all unique roles
        4. Get permissions for each role
        5. Fetch the permission catalogue
        6. Flag high-risk combinations

        Args:
            max_users: Max users to scan (0 = unlimited)
            filter_field: OData field to filter by (e.g. "department", "division",
                         "country", "location", "custom01"–"custom15")
            filter_value: Value to filter on (e.g. "Finance", "GBR")
        """
        errors: list[str] = []

        # Build the $filter expression
        base_filter = "status eq 'active' or status eq 't'"
        if filter_field and filter_value:
            escaped_val = odata_escape(filter_value)
            full_filter = f"{base_filter} and {filter_field} eq '{escaped_val}'"
        else:
            full_filter = base_filter

        # Step 1: Fetch users (optionally filtered)
        logger.info("Step 1/5: Fetching users...")
        if filter_field and filter_value:
            logger.info("  → Filter: %s = '%s'", filter_field, filter_value)
        try:
            users_raw = self._client.get(
                "User",
                select=["username", "userId", "firstName", "lastName", "status"],
                filter_expr=full_filter,
            )
            if max_users > 0:
                users_raw = users_raw[:max_users]
            logger.info("  → Found %d active users", len(users_raw))
        except SFClientError as exc:
            errors.append(f"Failed to fetch users: {exc}")
            users_raw = []

        # Step 2: Get role assignments for all users
        logger.info("Step 2/5: Fetching role assignments...")
        user_ids = [u.get("userId", u.get("username", "")) for u in users_raw if u.get("userId") or u.get("username")]

        user_role_map: dict[str, list[str]] = {}
        if user_ids:
            try:
                user_role_map = self.get_user_roles_report(user_ids)
            except SFClientError as exc:
                errors.append(f"Failed to fetch user role report: {exc}")

        # Check for RBP access denial early
        rbp_blocked = any(
            "403" in e or "RBP" in e or "Access denied" in e or "administrator" in e
            for e in errors
        )
        if not rbp_blocked and user_ids and not user_role_map:
            # try getPermissionMetadata as a probe — if it fails we know RBP is blocked
            try:
                _probe = self.get_permission_metadata()
            except SFClientError as exc:
                err_msg = str(exc)
                if "403" in err_msg or "administrator" in err_msg or "RBP" in err_msg:
                    errors.append(
                        "API user lacks RBP administrator access. "
                        "All RBP functions returned empty/blocked. "
                        "Grant via Admin Center > Manage Role-Based Permission Access"
                    )
                    rbp_blocked = True

        if rbp_blocked:
            logger.warning("RBP access blocked — returning partial report (users only)")
            return PermissionScanReport(
                roles=[],
                users=[
                    UserRoleAssignment(
                        user_id=u.get("userId", u.get("username", "")),
                        username=u.get("username", ""),
                        full_name=f"{u.get('firstName', '')} {u.get('lastName', '')}".strip(),
                        status=str(u.get("status", "active")),
                        is_inactive=str(u.get("status", "")).lower() in ("f", "inactive", "0"),
                    )
                    for u in users_raw
                ],
                catalogue=None,
                tenant_url=self._base_url,
                total_roles=0,
                total_users=len(users_raw),
                errors=errors,
            )

        # Step 3: Discover unique roles
        logger.info("Step 3/5: Discovering roles...")
        all_roles_set: set[str] = set()
        for roles in user_role_map.values():
            all_roles_set.update(roles)

        # Also try to find roles from user role IDs (numeric role IDs)
        all_role_ids: list[str] = sorted(all_roles_set)
        logger.info("  → Found %d unique roles", len(all_role_ids))

        # Step 4: Get permissions for each role
        logger.info("Step 4/5: Fetching role permissions...")
        role_permissions_map: dict[str, list[str]] = {}
        if all_role_ids:
            try:
                role_permissions_map = self.get_roles_permissions(all_role_ids)
            except SFClientError as exc:
                errors.append(f"Failed to fetch role permissions: {exc}")
                # Still add roles with empty perms
                for rid in all_role_ids:
                    role_permissions_map.setdefault(rid, [])

        # Step 5: Fetch permission catalogue
        logger.info("Step 5/5: Fetching permission catalogue...")
        catalogue: PermissionCatalogue | None = None
        try:
            catalogue = self.get_permission_metadata()
        except SFClientError as exc:
            errors.append(f"Failed to fetch permission metadata: {exc}")

        # ── Build data models ───────────────────────────────────────────

        # Build role objects
        role_objects: list[PermissionRole] = []
        for role_id in all_role_ids:
            perms = role_permissions_map.get(role_id, [])
            role_objects.append(
                PermissionRole(
                    role_id=role_id,
                    role_name=role_id,
                    permissions=perms,
                    is_empty=len(perms) == 0,
                )
            )

        # Build user objects
        users: list[UserRoleAssignment] = []
        for u in users_raw:
            uid = u.get("userId", u.get("username", ""))
            username = u.get("username", "")
            full_name = (
                f"{u.get('firstName', '')} {u.get('lastName', '')}".strip()
            )
            assigned_roles = user_role_map.get(uid, []) or user_role_map.get(username, [])
            status = u.get("status", "active")
            users.append(
                UserRoleAssignment(
                    user_id=uid,
                    username=username,
                    full_name=full_name,
                    status=str(status),
                    role_ids=assigned_roles,
                    role_names=assigned_roles,
                    is_inactive=str(status).lower() in ("f", "inactive", "0"),
                )
            )

        # Count users per role
        role_user_counts: dict[str, int] = {}
        for u in users:
            for rn in u.role_names:
                role_user_counts[rn] = role_user_counts.get(rn, 0) + 1
        for role_obj in role_objects:
            role_obj.user_count = role_user_counts.get(role_obj.role_name, 0)

        # Identify empty roles
        empty_roles = [r for r in role_objects if r.is_empty]

        # Identify high-risk roles
        high_risk_roles = self._flag_high_risk_roles(role_objects, catalogue)

        logger.info("Scan complete: %d roles, %d users", len(role_objects), len(users))
        if empty_roles:
            logger.info("  → %d roles have zero permissions", len(empty_roles))
        if high_risk_roles:
            logger.info("  → %d roles flagged high-risk", len(high_risk_roles))

        return PermissionScanReport(
            roles=role_objects,
            users=users,
            catalogue=catalogue,
            tenant_url=self._base_url,
            total_roles=len(role_objects),
            total_users=len(users),
            empty_roles=empty_roles,
            high_risk_roles=high_risk_roles,
            errors=errors,
        )

    def _flag_high_risk_roles(
        self,
        roles: list[PermissionRole],
        catalogue: PermissionCatalogue | None,
    ) -> list[tuple[str, str, list[str]]]:
        """Flag roles with dangerous permission combinations.

        Returns list of (role_name, risk_description, matching_perms).
        """
        flags: list[tuple[str, str, list[str]]] = []

        for role in roles:
            # Check for sensitive single permissions
            sensitive_found: list[str] = []
            for sp_code, sp_label in _SENSITIVE_PERMISSIONS.items():
                # Match against the raw permission code
                for rp in role.permissions:
                    if sp_code.lower() in rp.lower():
                        sensitive_found.append(f"{sp_label} ({rp})")
                        break

            # Check for dangerous combinations
            for combo in _HIGH_RISK_COMBOS:
                matched = [c for c in combo if any(c.lower() in rp.lower() for rp in role.permissions)]
                if len(matched) >= 2:
                    flags.append((
                        role.role_name,
                        f"Dangerous combo: {', '.join(_SENSITIVE_PERMISSIONS.get(m, m) for m in matched)}",
                        matched,
                    ))

            # Flag roles with many sensitive permissions
            if len(sensitive_found) >= 3:
                flags.append((
                    role.role_name,
                    f"Over-privileged: {len(sensitive_found)} sensitive permissions granted",
                    sensitive_found,
                ))

        return flags

    def build_permission_matrix(
        self, report: PermissionScanReport
    ) -> dict[str, Any]:
        """Build a permission matrix for Excel/HTML export.

        Returns:
            {
                "roles": [{role_name, permission_count, permissions: [...]}, ...],
                "users": [{user_id, roles: [...], effective_permissions: [...]}, ...],
                "risk_flags": [...],
                "catalogue": {...}
            }
        """
        # Resolve permission codes to labels via catalogue
        code_to_label: dict[str, str] = {}
        if report.catalogue:
            for cat_perms in report.catalogue.categories.values():
                for p in cat_perms:
                    code_to_label[p["code"]] = p["label"]

        matrix: dict[str, Any] = {
            "tenant_url": report.tenant_url,
            "scan_timestamp": __import__("datetime").datetime.now().isoformat(),
            "summary": {
                "total_roles": report.total_roles,
                "total_users": report.total_users,
                "empty_roles": len(report.empty_roles),
                "high_risk_roles": len(report.high_risk_roles),
                "errors": report.errors,
            },
            "roles": [],
            "users": [],
            "risk_flags": [
                {"role_name": r[0], "risk": r[1], "permissions": r[2]}
                for r in report.high_risk_roles
            ],
        }

        for role in report.roles:
            labeled_perms = [
                {"code": p, "label": code_to_label.get(p, p)}
                for p in role.permissions
            ]
            matrix["roles"].append({
                "role_id": role.role_id,
                "role_name": role.role_name,
                "permission_count": len(role.permissions),
                "user_count": role.user_count,
                "is_empty": role.is_empty,
                "permissions": labeled_perms,
            })

        # Get effective permissions for users (in batches)
        user_ids = [u.user_id for u in report.users if u.user_id]
        users_perms_map: dict[str, list[str]] = {}
        if user_ids:
            try:
                users_perms_map = self.get_users_permissions(user_ids)
            except SFClientError:
                logger.warning("Could not fetch user effective permissions")

        for user in report.users:
            effective_perms = users_perms_map.get(user.user_id, []) or users_perms_map.get(user.username, [])
            labeled_effective = [
                {"code": p, "label": code_to_label.get(p, p)}
                for p in effective_perms
            ]
            matrix["users"].append({
                "user_id": user.user_id,
                "username": user.username,
                "full_name": user.full_name,
                "status": user.status,
                "role_ids": user.role_ids,
                "role_names": user.role_names,
                "is_inactive": user.is_inactive,
                "effective_permissions": labeled_effective,
                "effective_perm_count": len(effective_perms),
            })

        return matrix

    def export_to_excel(
        self, report: PermissionScanReport, output_path: str
    ) -> None:
        """Export the scan report to an Excel workbook.

        Sheets:
        1. Summary — overview stats + risk flags
        2. Roles — each role with its permissions
        3. Users — each user with role assignments
        4. Empty Roles — roles with no permissions
        5. Permission Catalogue — full permission list
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: "
                "pip install openpyxl"
            ) from None

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        if ws_summary is None:
            ws_summary = wb.create_sheet("Summary", 0)
        assert ws_summary is not None

        # ── Styles ──
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="21446F", end_color="21446F", fill_type="solid")
        risk_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
        warn_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
        wrap = Alignment(wrap_text=True, vertical="top")

        def _write_header(ws, headers: list[str]) -> None:
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = wrap

        def _auto_width(ws, max_width: int = 60) -> None:
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                lengths = []
                for cell in col:
                    if cell.value:
                        lengths.append(min(len(str(cell.value)), max_width))
                if lengths:
                    ws.column_dimensions[col_letter].width = max(lengths) + 2

        # ── Sheet 1: Summary ──
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.cell(row=1, column=1, value="SF Permission Analyzer - Scan Report").font = Font(bold=True, size=14)
        ws_summary.cell(row=2, column=1, value=f"Tenant: {report.tenant_url}")
        ws_summary.cell(row=4, column=1, value="Metric").font = header_font
        ws_summary.cell(row=4, column=2, value="Value").font = header_font
        ws_summary.cell(row=4, column=1).fill = header_fill
        ws_summary.cell(row=4, column=2).fill = header_fill

        metrics = [
            ("Total Roles", report.total_roles),
            ("Total Users", report.total_users),
            ("Empty Roles (no permissions)", len(report.empty_roles)),
            ("High Risk Roles", len(report.high_risk_roles)),
            ("Errors", len(report.errors)),
        ]
        for i, (k, v) in enumerate(metrics, 5):
            ws_summary.cell(row=i, column=1, value=k)
            cell = ws_summary.cell(row=i, column=2, value=v)
            if k.startswith("High Risk") and v > 0:
                cell.fill = risk_fill
            if k.startswith("Empty") and v > 0:
                cell.fill = warn_fill

        # ── Sheet 2: Roles ──
        ws_roles = wb.create_sheet("Roles")
        _write_header(ws_roles, ["Role Name", "Role ID", "Permission Count", "User Count", "Permissions", "Is Empty"])
        for row_idx, role in enumerate(report.roles, 2):
            ws_roles.cell(row=row_idx, column=1, value=role.role_name)
            ws_roles.cell(row=row_idx, column=2, value=role.role_id)
            ws_roles.cell(row=row_idx, column=3, value=len(role.permissions))
            ws_roles.cell(row=row_idx, column=4, value=role.user_count)
            ws_roles.cell(row=row_idx, column=5, value=", ".join(role.permissions) if role.permissions else "(none)")
            ws_roles.cell(row=row_idx, column=6, value="Yes" if role.is_empty else "No")
            if role.is_empty:
                for c in range(1, 7):
                    ws_roles.cell(row=row_idx, column=c).fill = warn_fill
        _auto_width(ws_roles)

        # ── Sheet 3: Users ──
        ws_users = wb.create_sheet("Users")
        _write_header(ws_users, ["User ID", "Username", "Full Name", "Status", "Roles", "Role Count"])
        for row_idx, user in enumerate(report.users, 2):
            ws_users.cell(row=row_idx, column=1, value=user.user_id)
            ws_users.cell(row=row_idx, column=2, value=user.username)
            ws_users.cell(row=row_idx, column=3, value=user.full_name)
            ws_users.cell(row=row_idx, column=4, value=user.status)
            ws_users.cell(row=row_idx, column=5, value=", ".join(user.role_names) if user.role_names else "(none)")
            ws_users.cell(row=row_idx, column=6, value=len(user.role_names))
        _auto_width(ws_users)

        # ── Sheet 4: Empty Roles ──
        ws_empty = wb.create_sheet("Empty Roles")
        _write_header(ws_empty, ["Role Name", "Role ID"])
        for row_idx, role in enumerate(report.empty_roles, 2):
            ws_empty.cell(row=row_idx, column=1, value=role.role_name)
            ws_empty.cell(row=row_idx, column=2, value=role.role_id)
        _auto_width(ws_empty)

        # ── Sheet 5: Permission Catalogue ──
        ws_cat = wb.create_sheet("Permission Catalogue")
        _write_header(ws_cat, ["Category", "Permission Code", "Label"])
        row_idx = 2
        if report.catalogue:
            for cat_name, perms in report.catalogue.categories.items():
                for p in perms:
                    ws_cat.cell(row=row_idx, column=1, value=cat_name)
                    ws_cat.cell(row=row_idx, column=2, value=p["code"])
                    ws_cat.cell(row=row_idx, column=3, value=p["label"])
                    row_idx += 1
        _auto_width(ws_cat)

        # ── Sheet 6: Risk Flags ──
        ws_risk = wb.create_sheet("Risk Flags")
        _write_header(ws_risk, ["Role Name", "Risk Description", "Matching Permissions"])
        for row_idx, (role_name, risk_desc, perms) in enumerate(report.high_risk_roles, 2):
            ws_risk.cell(row=row_idx, column=1, value=role_name)
            ws_risk.cell(row=row_idx, column=2, value=risk_desc)
            ws_risk.cell(row=row_idx, column=3, value=", ".join(perms))
            for c in range(1, 4):
                ws_risk.cell(row=row_idx, column=c).fill = risk_fill
        _auto_width(ws_risk)

        wb.save(output_path)
        logger.info("Report exported to %s", output_path)
